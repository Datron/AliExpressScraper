from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import time


def get_url(choice):
    # choice 1 is best selling products weekly
    if choice == 1:
        url = "https://bestselling.aliexpress.com/en"
    # choice 2 is best selling products daily
    elif choice == 2:
        url = "https://sale.aliexpress.com/__pc/hot-products.htm"
    # choice 3 is latest products new on Ali Express
    elif choice == 3:
        url = "https://sale.aliexpress.com/__pc/newarrivals.htm"
    elif choice == 4:
        # categories in hot products
        url = "https://sale.aliexpress.com/__pc/bestselling.htm?spm=a2g01.8005310.2004.101.5d8fde54rH5A6M#701"
    return url


def get_top_products(url):
    global driver
    wb = openpyxl.Workbook()
    print('category', 'count', sep='\t\t\t')
    sheet = wb.active
    sheet.title = 'All'
    configure_sheet(sheet)
    driver.get(url)
    categories = {'Women': 'https://bestselling.aliexpress.com/en#701',
                  'Men': 'https://bestselling.aliexpress.com/en#702',
                  'Electronics': 'https://bestselling.aliexpress.com/en#703',
                  'Sports': 'https://bestselling.aliexpress.com/en#704',
                  'Health & Beauty': 'https://bestselling.aliexpress.com/en#705',
                  'Kids & Baby': 'https://bestselling.aliexpress.com/en#706',
                  'Home & Garden': 'https://bestselling.aliexpress.com/en#707',
                  'Automotive': 'https://bestselling.aliexpress.com/en#708'}
    top10 = driver.find_elements_by_class_name("top10-item")
    print('All', len(top10), sep='\t\t\t\t\t')
    total_products_count = 0
    href = []
    for item in top10:
        a = item.find_element_by_tag_name('a')
        href.append(a.get_attribute('href'))
    i = 2
    for link in href:
        scrape_product_page(link, sheet, i)
        i += 1
    wb.save('top_products.xlsx')
    return


def get_hot_products(url):
    global driver
    driver.get(url)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "All"
    configure_sheet(sheet)
    items = driver.find_elements_by_class_name('image-box')
    i = 2
    href = []
    for item in items:
        a = item.find_element_by_tag_name('a')
        href.append(a.get_attribute('href'))
    for link in href:
        scrape_product_page(link, sheet, i)
        i += 1
    driver.get(get_url(4))
    ul = driver.find_element_by_class_name("menuitems")
    links = ul.find_elements_by_tag_name("a")
    i = 0
    j = 350
    titles = []
    for link in links:
        try:
            # link.click()
            titles.append(link.text)
            driver.execute_script("window.scrollBy(" + str(i) + "," + str(j) + ");")
            time.sleep(7)
            element = WebDriverWait(driver, 20).until(lambda driver: driver.find_element_by_class_name("detail-box"))
            i = j
            j += 350
        except Exception as e:
            print(e)
    elements = driver.find_elements_by_class_name("detail-box")
    products = []
    for element in elements:
        link = element.find_element_by_tag_name('a')
        products.append(link.get_attribute('href'))
    for title in titles:
        j = 2
        sheet = wb.create_sheet(title)
        configure_sheet(sheet)
        for product in products:
            print("link:"+product)
            scrape_product_page(product, sheet, j)
            products.remove(product)
            print("length of products after removal:"+str(len(products)))
            j += 1
            if j == 22:
                break
    total = len(elements)+len(items)
    print("hot products", total, sep="\t\t")
    wb.save("hot_products.xlsx")
    return


def get_latest_products(url):
    global driver
    driver.get(url)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "All"
    configure_sheet(sheet)
    items = driver.find_elements_by_class_name('pro-img')
    print("latest products", len(items), sep='\t\t')
    i = 2
    href = []
    for item in items:
        a = item.find_element_by_tag_name('a')
        href.append(a.get_attribute('href'))
    for link in href:
        scrape_product_page(link, sheet, i)
        i += 1
    wb.save("latest_products.xlsx")
    return


def scrape_product_page(url, sheet, i):
    global driver
    driver.get(url)
    c = str(i)
    try:
        sheet['A' + c] = i - 1
    except Exception as e:
        k = 0
    try:
        sheet['B' + c] = driver.find_element_by_class_name('product-name').text
    except Exception as e:
        k = 0
    try:
        sheet['G' + c] = driver.find_element_by_class_name('percent-num').text
    except Exception as e:
        k = 0
    try:
        sheet['H' + c] = driver.find_element_by_class_name('rantings-num').text
    except Exception as e:
        k = 0
    try:
        sheet['I' + c] = driver.find_element_by_class_name('order-num').text
    except Exception as e:
        k = 0
    try:
        sheet['F' + c] = driver.find_element_by_class_name('p-symbol').text
    except Exception as e:
        k = 0
    try:
        sheet['C' + c] = driver.find_element_by_id('j-sku-price').text
    except Exception as e:
        k = 0
    try:
        sheet['D' + c] = driver.find_element_by_id('j-sku-discount-price').text
    except Exception as e:
        k = 0
    try:
        sheet['E' + c] = driver.find_element_by_id('j-product-shipping').text
    except Exception as e:
        k = 0
    try:
        sheet['J' + c] = driver.find_element_by_id('p-available-stock').text
    except Exception as e:
        k = 0
    return


def configure_sheet(sheet):
    sheet['A1'] = 'SL.NO'
    sheet['B1'] = 'Product Title'
    sheet['C1'] = 'Product Actual Price'
    sheet['D1'] = 'Product Discount Price'
    sheet['E1'] = 'Shipping cost'
    sheet['F1'] = 'Currency'
    sheet['G1'] = 'Rating'
    sheet['H1'] = 'Number of Ratings'
    sheet['I1'] = 'Number of products Ordered'
    sheet['J1'] = 'Stock left'
    return


print("parsing....")
options = Options()
options.add_argument("--headless")
driver = webdriver.Firefox(firefox_options=options, executable_path="geckodriver.exe")
# print("-------------------TOP PRODUCTS------------------------")
# get_top_products(get_url(1))
# print("-------------------HOT PRODUCTS------------------------")
get_hot_products(get_url(2))
# print("-------------------LATEST PRODUCTS------------------------")
# get_latest_products(get_url(3))
driver.close()
